'''
MIT License

Copyright 2024 National Technology & Engineering Solutions of Sandia, LLC (NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S. Government retains certain rights in this software.

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
'''

import openpyxl
import json
import sys
import os
import re

def create_directory(xlsx_file):
    dir_name = xlsx_file.split('.')[0]
    path = './' + dir_name

    if not os.path.exists(path):
        os.mkdir(path)
        print(">>> Folder %s created!" % path)
        return path
    else:
        print(">>> Folder %s already exists" % path)
        return 'Error'


def create_json_files(xlsx_file):
    certification = openpyxl.load_workbook(xlsx_file)
    models = certification.sheetnames

    path = create_directory(xlsx_file)

    if path != 'Error':
        for model in models:
            if model not in ['Instructions', 'Testing Info', 'Certification Info']:
                json_data = {}

                for row in range(1, certification[model].max_row + 1):
                    json_data[certification[model].cell(row, 1).value] = []
                    for col in range(2, certification[model].max_column + 1):
                        json_data[certification[model].cell(row, 1).value].append(certification[model].cell(row, col).value)
                
                with open(f"{path}/raw_json_{model}.json", "w") as f:
                    json.dump(json_data,f)

        generate_device_map(path)
    else:
        print('>>> There is a path error. Exiting...')

def extract_text(text):
    if text is not None:
        match = re.search(r'\((.*?)\)', text)
        if match:
            return match.group(1)
        else:
            return text
    else:
        return None

def generate_device_map(directory):
    raw_json_files = [file for file in os.listdir(directory) if file.endswith('json')]

    models_data = []
    for raw_json_file in raw_json_files:
        with open(directory + '/' + raw_json_file) as json_file:
            model_data = json.load(json_file)
            if model_data:
                models_data.append({})
                for rtg, value in model_data.items():
                    if rtg != 'Address':
                        if extract_text(value[0]):
                            models_data[-1][extract_text(value[0])] = value[3]
    
    output_model_data = {'models': models_data}
    with open(f"{directory}/device_map.json", "w") as f:
        json.dump(output_model_data, f, indent=4)
    
    print('>>> DER Device map is successfully generated.')


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('>>> Please provide exactly one argument - the Excel file to be parsed for the DER device.')
    else:
        xlsx_file = sys.argv[1]
        create_json_files(xlsx_file)
